#!/usr/bin/env python3
"""
build_data.py — erzeugt data.js aus dem Master-Dokument tdf_tippspiel_master.xlsx.

Die xlsx ist die einzige Datenquelle ("single source of truth"). Nach jeder
Aktualisierung (neue Etappe eingetragen, Ergebnis korrigiert, Spruch geändert …)
einfach dieses Skript ausführen:

    python build_data.py

index.html / support.js bleiben unverändert — die Export-Struktur von data.js
ist identisch zu vorher, nur dass die Etappenwerte jetzt echt (aus den
Ergebnissen berechnet) statt geschätzt sind.

Wie die Punkte entstehen (identisch zur Excel-Logik, hier in Python nachgebaut,
damit das Ergebnis NICHT davon abhängt, ob Excel die Formeln neu berechnet hat):

  Punkte eines Fahrers/Teams je Etappe
      = Summe der Punktematrix-Werte für jede Platzierung/Wertung,
        in der er/sie in dieser Etappe in 'Ergebnisse_Roh' steht.
  Punkte eines Spielers je Etappe
      = Summe der Punkte all seiner getippten Fahrer + seines Team-Tipps.

Roh-Eingaben, die gepflegt werden (alles andere wird berechnet):
  - "Ergebnisse_Roh"  Platzierungen 1–15 + Trikots + Mannschaft je Etappe
  - "Punktematrix"     Punkte je Platzierung / Wertung
  - "Spieler_Teams"    Spalte A/B: welcher Spieler tippt welche Fahrer + 1 Team
  - "Startlist2026"    Fahrer-Stammdaten (BIB, Specialty, Age, Rider, Team)
  - "Quotes"           Teamfunk / Trash-Talk (Spalten: Spieler, Spruch)
  - "RIderPoints.csv"  Fahrer-/Team-Kosten (Budget fuers Kanter-Kader, s.u.)

Kanter-Kader: das budget-optimale Team, das man mit dem Wissen bis zu einer
Etappe haette zusammenstellen koennen (<=560 Kosten aus RIderPoints.csv,
<=15 Fahrer + hoechstens 1 optionaler Team-Tipp). Wird hier als 0/1-Rucksack
geloest (Portierung von optimal_team()/optimal_team_full() aus index.qmd).
"""

import csv
import itertools
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt — bitte 'pip install openpyxl' ausführen.")

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "tdf_tippspiel_master.xlsx"
COSTS_CSV = ROOT / "RIderPoints.csv"
OUT = ROOT / "data.js"
PHOTO_CACHE_FILE = ROOT / ".rider_photos_cache.json"
QUOTES_FILE = ROOT / "quotes.json"
WIKIPEDIA_API = "https://en.wikipedia.org/w/api.php"

# Gesamtzahl Etappen der Tour. Wie viele bereits gefahren sind, ergibt sich
# automatisch aus den befüllten Etappen-Spalten in 'Ergebnisse_Roh'.
STAGES_TOTAL = 21


# --------------------------------------------------------------------------
# Hilfsfunktionen
# --------------------------------------------------------------------------
_REF_RE = re.compile(r"^=\s*(?:'([^']+)'|([A-Za-z0-9_]+))?!?\$?([A-Z]+)\$?(\d+)\s*$")


def resolve(wb, value, depth=0):
    """Löst einfache Zellformeln wie =Startlist2026!D2 oder =A1 zum Wert auf.
    Nötig, weil einige Tipp-Zellen als Verweise statt als Text gespeichert sind
    (und deren Cache-Wert beim Bearbeiten verloren gehen kann). Komplexe
    Formeln -> None."""
    if not isinstance(value, str) or not value.startswith("="):
        return value
    if depth > 5:
        return None
    m = _REF_RE.match(value)
    if not m:
        return None
    sheet = m.group(1) or m.group(2)
    ref = f"{m.group(3)}{m.group(4)}"
    try:
        ws = wb[sheet] if sheet else None
    except KeyError:
        return None
    if ws is None:
        return None
    return resolve(wb, ws[ref].value, depth + 1)


def split_rider_name(core):
    """'PogačarTadej' -> 'Pogačar Tadej'. Leerzeichen an der letzten
    Kleinbuchstabe→Großbuchstabe-Grenze (Nachname/Vorname)."""
    boundary = -1
    for i in range(1, len(core)):
        if core[i - 1].islower() and core[i].isupper():
            boundary = i
    return core if boundary < 1 else core[:boundary] + " " + core[boundary:]


def js_str(s):
    return "'" + str(s).replace("\\", "\\\\").replace("'", "\\'") + "'"


def stage_columns(ws, header_row=1):
    """Spaltenindizes der Etappen-Spalten (S01, 'S 01', Stage01 …)."""
    cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and re.match(r"^\s*(S|Stage)\s*0*\d+\s*$", v):
            cols.append(c)
    return cols


# --------------------------------------------------------------------------
# Fahrerfotos: Wikipedia-"Page Image" (dieselbe Vorschau, die auch Infobox/
# Linkvorschau nutzt), gefunden per Volltextsuche statt exaktem Titel — unsere
# Fahrernamen stehen als "Nachname Vorname" (siehe split_rider_name), waehrend
# Wikipedia-Artikel unter "Vorname Nachname" laufen; die Suche ist tolerant
# gegenueber der Reihenfolge. Mit lokalem Cache, damit nicht bei jedem Build
# erneut angefragt wird (und es auch offline funktioniert, sobald gefuellt).
# Kein/falscher Treffer -> None, dann zeigt das Dashboard einen Initialen-
# Avatar statt eines Fotos.
# --------------------------------------------------------------------------
def _http_get_json(url, timeout=8):
    req = urllib.request.Request(url, headers={"User-Agent": "TdF2026-Tipprunde/1.0 (build_data.py)"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


class _PhotoFetchError(Exception):
    """Netzwerk-/API-Fehler beim Foto-Abruf — im Unterschied zu einem
    bestaetigten 'kein Foto gefunden' NICHT cachen, damit der naechste
    Build es erneut versucht statt den Fahrer dauerhaft ohne Foto zu lassen."""


def _wikipedia_photo(name, size=200):
    """Bester Volltext-Treffer fuer `name` + dessen Page-Image-Thumbnail,
    in einer einzigen API-Anfrage (generator=search + prop=pageimages).
    None = bestaetigt kein Foto gefunden. Wirft _PhotoFetchError bei
    Netzwerk-/API-Problemen (siehe fetch_rider_photos)."""
    params = {
        "action": "query", "generator": "search", "gsrsearch": name, "gsrlimit": "1",
        "prop": "pageimages", "piprop": "thumbnail", "pithumbsize": str(size), "format": "json",
    }
    url = WIKIPEDIA_API + "?" + urllib.parse.urlencode(params)
    data = None
    for attempt in range(3):
        try:
            data = _http_get_json(url)
            break
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < 2:
                time.sleep(2 + attempt * 2)
                continue
            raise _PhotoFetchError(str(e)) from e
        except (urllib.error.URLError, TimeoutError, ValueError, OSError) as e:
            raise _PhotoFetchError(str(e)) from e
    if data is None:
        raise _PhotoFetchError("retries exhausted")
    for page in data.get("query", {}).get("pages", {}).values():
        thumb = page.get("thumbnail", {}).get("source")
        if thumb:
            return thumb
    return None


def load_photo_cache():
    if PHOTO_CACHE_FILE.exists():
        try:
            return json.loads(PHOTO_CACHE_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            return {}
    return {}


def fetch_rider_photos(names, skip=False):
    """dict Fahrername -> Wikipedia-Foto-URL oder None. Nutzt/pflegt
    PHOTO_CACHE_FILE; fragt nur Namen an, die noch nicht im Cache stehen."""
    cache = load_photo_cache()
    if skip:
        return cache
    missing = [n for n in dict.fromkeys(names) if n not in cache]
    if not missing:
        return cache
    changed = False
    for name in missing:
        try:
            cache[name] = _wikipedia_photo(name)
            changed = True
        except _PhotoFetchError:
            pass  # nicht cachen -> naechster Build versucht's erneut
        time.sleep(0.6)  # rücksichtsvoll gegenüber der Wikipedia-API
    if changed:
        PHOTO_CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=1, sort_keys=True), encoding="utf-8")
    return cache


_LATIN_ASCII_EXTRA = str.maketrans({
    "ø": "o", "Ø": "O", "æ": "ae", "Æ": "AE", "ł": "l", "Ł": "L",
    "đ": "d", "Đ": "D", "ß": "ss", "þ": "th", "Þ": "Th", "ð": "d", "Ð": "D",
})


def match_key(s):
    """Namensschlüssel für robusten Abgleich zwischen RIderPoints.csv und
    Startlist2026: nur Kleinbuchstaben, Diakritika/Leerzeichen/Zeichensetzung
    entfernt (identisch zu match_key()/stri_trans_general(...,'Latin-ASCII')
    in index.qmd). unicodedata deckt reine Akzente ab (é, ö, č …); die paar
    Buchstaben, die nicht per Combining-Mark zerlegbar sind (ø, æ, ł, ß …),
    werden vorher explizit ersetzt."""
    s = str(s).translate(_LATIN_ASCII_EXTRA)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z]", "", s.lower())


def parse_rider_prices(path):
    """RIderPoints.csv ist ein Excel-Rohexport: 5 Team-Blöcke nebeneinander
    (je 4 Spalten: Label/Name/Rating/Leerspalte), jeder Block gefolgt von 8
    Fahrer-Zeilen, Blöcke durch Leerzeilen getrennt. Portierung von
    parse_rider_prices() in index.qmd. Rückgabe: (rider_prices, team_prices),
    beide dict match_key -> Rating (int)."""
    if not path.exists():
        return {}, {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))[1:]  # erste Zeile: nur "Rating"-Spaltenköpfe je Block

    rider_prices, team_prices = {}, {}
    i, n = 0, len(rows)
    while i < n:
        row = rows[i]
        if all(not cell.strip() for cell in row):
            i += 1
            continue
        team_blocks = {}  # Block-Start-Spalte -> Teamname
        for bs in range(0, len(row), 4):
            label = row[bs].strip() if bs < len(row) else ""
            name = row[bs + 1].strip() if bs + 1 < len(row) else ""
            rating = row[bs + 2].strip() if bs + 2 < len(row) else ""
            if label == "Team" and name:
                team_blocks[bs] = name
                try:
                    team_prices[match_key(name)] = int(float(rating))
                except ValueError:
                    pass
        if team_blocks:
            i += 1
            for _ in range(8):
                if i >= n:
                    break
                row2 = rows[i]
                for bs in team_blocks:
                    name = row2[bs + 1].strip() if bs + 1 < len(row2) else ""
                    rating = row2[bs + 2].strip() if bs + 2 < len(row2) else ""
                    if name:
                        try:
                            rider_prices[match_key(name)] = int(float(rating))
                        except ValueError:
                            pass
                i += 1
        else:
            i += 1
    return rider_prices, team_prices


# --------------------------------------------------------------------------
# Kanter-Kader: 0/1-Rucksack fürs budget-optimale Team
# --------------------------------------------------------------------------
def optimal_team(pool, k=15, budget=560):
    """Wählt aus `pool` (Liste von dicts mit rating/points) bis zu `k`
    Einträge (nicht zwingend alle) mit Kostensumme <= `budget` (Rohwert,
    wird intern in 10er-Einheiten gerechnet), die die Punktesumme
    maximieren. Portierung von optimal_team() in index.qmd."""
    costs = [round(item["rating"] / 10) for item in pool]
    values = [item["points"] for item in pool]
    budget_u = budget // 10
    n = len(pool)
    NEG = float("-inf")
    dp = [[NEG] * (budget_u + 1) for _ in range(k + 1)]
    dp[0] = [0] * (budget_u + 1)
    chosen = [[[False] * (budget_u + 1) for _ in range(k + 1)] for _ in range(n)]
    for idx in range(n):
        c_i, v_i = costs[idx], values[idx]
        if c_i <= 0 or c_i > budget_u:
            continue
        for kk in range(k, 0, -1):
            for b in range(budget_u, c_i - 1, -1):
                candidate = dp[kk - 1][b - c_i] + v_i
                if candidate > dp[kk][b]:
                    dp[kk][b] = candidate
                    chosen[idx][kk][b] = True
    best_value, best_kk, best_b = 0, 0, 0
    for kk in range(k + 1):
        for b in range(budget_u + 1):
            if dp[kk][b] > best_value:
                best_value, best_kk, best_b = dp[kk][b], kk, b
    selected = []
    kk, b = best_kk, best_b
    for idx in range(n - 1, -1, -1):
        if kk > 0 and chosen[idx][kk][b]:
            selected.append(idx)
            kk -= 1
            b -= costs[idx]
    riders = sorted((pool[i] for i in selected), key=lambda r: -r["points"])
    total_cost = sum(pool[i]["rating"] for i in selected)
    return riders, best_value, total_cost


def optimal_team_full(pool_riders, team_options, k=15, budget=560):
    """Wie ein echter Tipp: bis zu 15 Fahrer + höchstens 1 Team-Tipp (oder
    keiner — eine 'kein Team'-Option mit 0 Kosten/0 Punkten steht immer zur
    Wahl), Gesamtkosten <= budget. Probiert jede bekannte Teamoption durch
    und behält die beste Gesamtpunktzahl. Portierung von
    optimal_team_full() in index.qmd."""
    options = [{"team": None, "rating": 0, "points": 0}] + team_options
    best = None
    for opt in options:
        team_rating = opt["rating"]
        if team_rating is None or team_rating > budget:
            continue
        riders, riders_value, riders_cost = optimal_team(pool_riders, k=k, budget=budget - team_rating)
        total = riders_value + opt["points"]
        if best is None or total > best["total"]:
            best = {
                "riders": riders, "riders_cost": riders_cost,
                "team": opt["team"], "team_rating": team_rating, "team_points": opt["points"],
                "total": total,
            }
    return best


# --------------------------------------------------------------------------
# Einlesen der Roh-Eingaben
# --------------------------------------------------------------------------
def read_points_matrix(wb):
    ws = wb["Punktematrix"]
    pts = {}
    for r in range(2, ws.max_row + 1):
        label, p = ws.cell(r, 1).value, ws.cell(r, 2).value
        if label is not None and p is not None:
            pts[str(label).strip()] = int(p)
    return pts


def read_entity_stage_points(wb, pts_map):
    """Punkte je Fahrer/Team und Etappe aus 'Ergebnisse_Roh'.
    Rückgabe: (dict entity -> [pkt je etappe], stages_done)."""
    ws = wb["Ergebnisse_Roh"]
    scols = stage_columns(ws)
    n = len(scols)
    ent = defaultdict(lambda: [0] * n)
    stage_has_data = [False] * n
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 2).value          # Platz / Trikot / Mannschaft
        if label is None:
            continue
        p = pts_map.get(str(label).strip())
        if p is None:
            continue
        for si, c in enumerate(scols):
            who = ws.cell(r, c).value
            if who not in (None, ""):
                ent[who][si] += p
                stage_has_data[si] = True
    stages_done = sum(1 for x in stage_has_data if x)
    return ent, stages_done


def read_tips(wb, teams):
    """Spalte A/B aus 'Spieler_Teams': Fahrer-Tipps + Team-Tipp je Spieler.
    Spalte B kann Verweise (=Startlist2026!Dxx) enthalten -> auflösen."""
    ws = wb["Spieler_Teams"]
    rider_picks = defaultdict(list)   # spieler -> [rider_key]
    team_pick = {}                    # spieler -> teamname
    tipped_by = defaultdict(set)      # rider_key -> {spieler}
    for r in range(2, ws.max_row + 1):
        sp = resolve(wb, ws.cell(r, 1).value)
        fa = resolve(wb, ws.cell(r, 2).value)
        if not sp or not fa:
            continue
        if fa in teams:
            team_pick[sp] = fa
        else:
            rider_picks[sp].append(fa)
            tipped_by[fa].add(sp)
    return rider_picks, team_pick, tipped_by


def read_startlist(wb):
    """Fahrer-Stammdaten (ohne Punkte — die werden berechnet)."""
    ws = wb["Startlist2026"]
    riders = {}
    teams = set()
    for r in range(2, ws.max_row + 1):
        bib = ws.cell(r, 1).value
        spec = ws.cell(r, 2).value
        age = ws.cell(r, 3).value
        key = ws.cell(r, 4).value
        team = ws.cell(r, 5).value
        if team:
            teams.add(team)
        if not key or not team or key == team:
            continue
        core = key[: -len(team)] if key.endswith(team) else key
        riders[key] = {
            "name": split_rider_name(core),
            "team": team,
            "spec": spec,
            "age": int(age) if age is not None else None,
            "bib": int(bib) if bib is not None else None,
        }
    return riders, teams


def read_quotes():
    """Teamfunk/Trash-Talk-Sprüche: bewusst NICHT in der xlsx, sondern in
    quotes.json — die xlsx wird unabhängig (z.B. von main) aktualisiert und
    hat/kennt dieses Sheet nicht; als eigene Datei überlebt das Feature jedes
    xlsx-Update unbeschadet."""
    if not QUOTES_FILE.exists():
        return []
    try:
        data = json.loads(QUOTES_FILE.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return []
    return [
        {"by": str(q["by"]), "txt": str(q["txt"])}
        for q in data
        if q.get("by") and q.get("txt")
    ]


# --------------------------------------------------------------------------
# Statistik ableiten
# --------------------------------------------------------------------------
def best_worst(stages):
    best_i = max(range(len(stages)), key=lambda i: (stages[i], -i))
    worst_i = min(range(len(stages)), key=lambda i: (stages[i], i))
    return [best_i + 1, stages[best_i]], [worst_i + 1, stages[worst_i]]


def compute_ranks(player_stages, names):
    """ranks[i] = Platz in der Gesamtwertung nach Etappe i.
    Standard-Competition-Ranking: Gleichstände teilen sich den Platz
    (…8, 8, 10…), wie in der ursprünglichen Kicktipp-Wertung."""
    n = len(next(iter(player_stages.values())))
    cum = {name: 0 for name in names}
    ranks = {name: [] for name in names}
    for i in range(n):
        for name in names:
            cum[name] += player_stages[name][i]
        for name in names:
            better = sum(1 for o in names if cum[o] > cum[name])
            ranks[name].append(better + 1)
    return ranks


# --------------------------------------------------------------------------
# Aufbau + Ausgabe
# --------------------------------------------------------------------------
def build():
    if not XLSX.exists():
        sys.exit(f"Master-Datei nicht gefunden: {XLSX}")
    # data_only=False: Literale kommen direkt, Formel-Verweise als Text
    # (werden via resolve() aufgelöst) — unabhängig von Excel-Cache-Werten.
    wb = openpyxl.load_workbook(XLSX, data_only=False)

    pts_map = read_points_matrix(wb)
    ent, stages_done = read_entity_stage_points(wb, pts_map)
    riders_meta, teams = read_startlist(wb)
    rider_picks, team_pick, tipped_by = read_tips(wb, teams)
    quotes = read_quotes()

    if stages_done == 0:
        sys.exit("Keine gefahrenen Etappen in 'Ergebnisse_Roh' gefunden.")

    def ent_stages(name):
        return (ent.get(name) or [0] * stages_done)[:stages_done]

    # ---- Kanter-Kader: Fahrer-/Team-Kosten aus RIderPoints.csv, per
    # Namensschlüssel (match_key) auf die echten Startlist2026-Namen gemappt.
    # Fahrer/Teams ohne Preis (z.B. Nachnominierungen) bleiben ohne Kosten.
    rider_prices, team_prices = parse_rider_prices(COSTS_CSV)

    rider_cost = {}
    for key, meta in riders_meta.items():
        team = meta["team"] or ""
        core = key[: -len(team)] if team and key.endswith(team) else key
        mk = match_key(core)
        if mk in rider_prices:
            rider_cost[key] = rider_prices[mk]

    team_cost = {}
    for team_name in teams:
        mk = match_key(team_name)
        if mk in team_prices:
            team_cost[team_name] = team_prices[mk]

    team_stage_points = {t: ent_stages(t) for t in teams}

    # ---- Spieler ----
    player_names = sorted(rider_picks.keys() | team_pick.keys(), key=str.casefold)
    player_stages = {}
    for name in player_names:
        st = [0] * stages_done
        for f in rider_picks.get(name, []):
            for i, v in enumerate(ent_stages(f)):
                st[i] += v
        tp = team_pick.get(name)
        if tp:
            for i, v in enumerate(ent_stages(tp)):
                st[i] += v
        player_stages[name] = st

    ranks = compute_ranks(player_stages, player_names)

    players_js = []
    for name in player_names:
        st = player_stages[name]
        total = sum(st)
        best, worst = best_worst(st)
        tp = team_pick.get(name)
        tip_pts = sum(ent_stages(tp)) if tp else 0
        players_js.append({
            "n": name,
            "avg": round(total / stages_done, 1),
            "best": best, "worst": worst,
            "tip": tp, "tipPts": tip_pts,
            "stages": st, "ranks": ranks[name],
        })

    # ---- Fahrer (nur getippte oder punktende) ----
    rel = []
    all_rider_keys = set(riders_meta) | set(tipped_by)
    for key in all_rider_keys:
        if key in teams:
            continue
        st = ent_stages(key)
        p = sum(st)
        by = sorted(tipped_by.get(key, ()), key=str.casefold)
        if p <= 0 and not by:
            continue
        meta = riders_meta.get(key)
        if meta is None:  # getippt, aber nicht in Startliste
            core = key
            meta = {"name": split_rider_name(core), "team": None,
                    "spec": None, "age": None, "bib": None}
        mx = max(st) if st else 0
        bd = [st.index(mx) + 1, mx] if mx > 0 else None
        rel.append({
            "n": meta["name"], "t": meta["team"], "s": meta["spec"],
            "a": meta["age"], "b": meta["bib"], "p": p, "bd": bd, "by": by,
            "st": st, "c": rider_cost.get(key),
        })
    rel.sort(key=lambda x: (-x["p"], x["n"].casefold()))

    # ---- Fahrerfotos (Wikipedia), gecacht -- SKIP_PHOTOS=1 zum Ueberspringen ----
    photo_cache = fetch_rider_photos([r["n"] for r in rel], skip=os.environ.get("SKIP_PHOTOS") == "1")
    for r in rel:
        r["img"] = photo_cache.get(r["n"])

    # ---- Team-Tipp-Punkte (nur getippte Teams mit Punkten) ----
    team_pts = {}
    for name in player_names:
        t = team_pick.get(name)
        if t:
            pts = sum(ent_stages(t))
            if pts > 0:
                team_pts[t] = pts

    # ---- Kanter-Kader: für jede Etappe unabhängig das budget-optimale Team
    # mit Kenntnis NUR der Ergebnisse bis zu dieser Etappe ("hätte man's bis
    # dahin gewusst") — Etappe 8 kann daher ein anderes Team zeigen als
    # Etappe 9. Für GLOBALE Vergleiche (Baseline im Punkteverlauf-Chart) wäre
    # es irreführend, diese wechselnden Teams zu einer Linie zusammenzu-
    # rechnen — stattdessen wird das für die letzte Etappe optimierte Team
    # als EIN festes Team verwendet (KANTER_KADER), dessen eigene kumulierte
    # Punkte über alle Etappen nachverfolgt werden.
    priced_riders = [
        {
            "key": key, "name": riders_meta[key]["name"], "team": riders_meta[key]["team"],
            "rating": rating, "stages": ent_stages(key),
        }
        for key, rating in rider_cost.items()
    ]
    priced_teams = [
        {"team": name, "rating": rating, "stages": team_stage_points[name]}
        for name, rating in team_cost.items()
    ]

    optimal_teams = []
    for stage_idx in range(stages_done):
        pool = [
            {"key": r["key"], "name": r["name"], "team": r["team"], "rating": r["rating"],
             "points": sum(r["stages"][: stage_idx + 1])}
            for r in priced_riders
        ]
        team_opts = [
            {"team": t["team"], "rating": t["rating"], "points": sum(t["stages"][: stage_idx + 1])}
            for t in priced_teams
        ]
        best = optimal_team_full(pool, team_opts)
        optimal_teams.append({
            "stage": stage_idx + 1,
            "team": best["team"], "teamCost": best["team_rating"], "teamPoints": best["team_points"],
            "totalCost": best["riders_cost"] + best["team_rating"], "totalPoints": best["total"],
            "riders": [
                {"key": r["key"], "n": r["name"], "t": r["team"], "c": r["rating"], "p": r["points"]}
                for r in best["riders"]
            ],
        })

    kanter_kader = None
    if optimal_teams:
        fixed = optimal_teams[-1]
        fixed_rider_keys = [r["key"] for r in fixed["riders"]]
        rider_cum = {k: list(itertools.accumulate(ent_stages(k))) for k in fixed_rider_keys}
        team_cum = (
            list(itertools.accumulate(team_stage_points[fixed["team"]]))
            if fixed["team"] else [0] * stages_done
        )
        cum = [sum(rider_cum[k][i] for k in fixed_rider_keys) + team_cum[i] for i in range(stages_done)]
        kanter_kader = {**fixed, "cum": cum}

    write_data_js(
        stages_done, players_js, rel, team_pts, quotes, team_cost, optimal_teams, kanter_kader,
        {t: team_stage_points[t] for t in team_cost},
    )

    print(f"✓ data.js geschrieben ({OUT})")
    print(f"  Etappen gefahren: {stages_done} / {STAGES_TOTAL}")
    print(f"  Spieler:          {len(players_js)}")
    print(f"  Fahrer (relevant):{len(rel)}")
    print(f"  Team-Tipps:       {len(team_pts)}")
    print(f"  Sprüche:          {len(quotes)}")


def write_data_js(
    stages_done, players_js, rel, team_pts, quotes, team_cost, optimal_teams, kanter_kader, team_stage_points,
):
    o = []
    o.append("// Tour de France 2026 Tipprunde — Daten")
    o.append("// ⚠️  AUTOMATISCH GENERIERT von build_data.py aus tdf_tippspiel_master.xlsx.")
    o.append("//     Nicht von Hand bearbeiten — stattdessen die xlsx ändern und")
    o.append("//     `python build_data.py` ausführen.")
    o.append(f"export const STAGES_DONE = {stages_done};")
    o.append(f"export const STAGES_TOTAL = {STAGES_TOTAL};")
    o.append("")

    o.append("export const PLAYERS = [")
    for p in players_js:
        tip = js_str(p["tip"]) if p["tip"] else "null"
        o.append(
            "  {{n:{n}, avg:{avg}, best:[{b0},{b1}], worst:[{w0},{w1}], "
            "tip:{tip}, tipPts:{tp}, stages:[{st}], ranks:[{rk}]}},".format(
                n=js_str(p["n"]), avg=p["avg"],
                b0=p["best"][0], b1=p["best"][1],
                w0=p["worst"][0], w1=p["worst"][1],
                tip=tip, tp=p["tipPts"],
                st=",".join(map(str, p["stages"])),
                rk=",".join(map(str, p["ranks"])),
            )
        )
    o.append("];")
    o.append("")

    o.append("// Fahrer: n, t=Team, s=Spezialist, a=Alter, b=Startnummer, p=Punkte gesamt,")
    o.append("// bd=[Etappe,Punkte] bester Tag, by=getippt von, st=Punkte je Etappe,")
    o.append("// c=Kosten aus RIderPoints.csv fürs Kanter-Kader (null wenn unbekannt),")
    o.append("// img=Wikipedia-Fotolink (null wenn kein Artikel/Foto gefunden)")
    o.append("export const RIDERS = [")
    for r in rel:
        bd = "null" if r["bd"] is None else f"[{r['bd'][0]},{r['bd'][1]}]"
        by = ",".join(js_str(x) for x in r["by"])
        t = js_str(r["t"]) if r["t"] else "null"
        s = js_str(r["s"]) if r["s"] else "null"
        a = r["a"] if r["a"] is not None else "null"
        b = r["b"] if r["b"] is not None else "null"
        st = ",".join(map(str, r["st"]))
        c = r["c"] if r["c"] is not None else "null"
        img = js_str(r["img"]) if r.get("img") else "null"
        o.append(
            "  {{n:{n}, t:{t}, s:{s}, a:{a}, b:{b}, p:{p}, bd:{bd}, by:[{by}], st:[{st}], c:{c}, img:{img}}},".format(
                n=js_str(r["n"]), t=t, s=s, a=a, b=b, p=r["p"], bd=bd, by=by, st=st, c=c, img=img,
            )
        )
    o.append("];")
    o.append("")

    o.append("// Teamwertungs-Punkte je Team (nur Teams mit Punkten)")
    o.append("export const TEAM_TIP_PTS = {")
    for t, pts in sorted(team_pts.items(), key=lambda kv: (-kv[1], kv[0])):
        o.append(f"  {js_str(t)}: {pts},")
    o.append("};")
    o.append("")

    o.append("// Team-Kosten aus RIderPoints.csv fürs Kanter-Kader (nur Teams mit bekanntem Preis)")
    o.append("export const TEAM_COSTS = {")
    for t, c in sorted(team_cost.items(), key=lambda kv: (-kv[1], kv[0])):
        o.append(f"  {js_str(t)}: {c},")
    o.append("};")
    o.append("")

    o.append("// Team-Punkte je Etappe (nur Teams mit bekanntem Preis) — Basis, um den")
    o.append("// Tagesbeitrag eines Team-Tipps im Kanter-Kader auszurechnen.")
    o.append("export const TEAM_STAGE_POINTS = {")
    for t in sorted(team_stage_points, key=str.casefold):
        o.append(f"  {js_str(t)}: [{','.join(map(str, team_stage_points[t]))}],")
    o.append("};")
    o.append("")

    def riders_js_array(riders):
        return ",".join(
            "{{n:{n}, t:{t}, c:{c}, p:{p}}}".format(
                n=js_str(r["n"]), t=js_str(r["t"]) if r["t"] else "null", c=r["c"], p=r["p"],
            )
            for r in riders
        )

    o.append("// Kanter-Kader: budget-optimales Team je Etappe (<=560 Kosten, <=15 Fahrer")
    o.append("// + optionaler Team-Tipp), unabhängig für jede Etappe optimiert — nur mit")
    o.append("// Wissen der Ergebnisse bis zu dieser Etappe ('hätte man's bis dahin gewusst').")
    o.append("export const OPTIMAL_TEAMS = [")
    for ot in optimal_teams:
        team_js = js_str(ot["team"]) if ot["team"] else "null"
        o.append(
            "  {{stage:{stage}, team:{team}, teamCost:{tc}, teamPoints:{tp}, totalCost:{totc}, "
            "totalPoints:{totp}, riders:[{riders}]}},".format(
                stage=ot["stage"], team=team_js, tc=ot["teamCost"], tp=ot["teamPoints"],
                totc=ot["totalCost"], totp=ot["totalPoints"], riders=riders_js_array(ot["riders"]),
            )
        )
    o.append("];")
    o.append("")

    o.append("// Festes Kanter-Kader-Team (Optimum der letzten gefahrenen Etappe) als")
    o.append("// durchgehende Baseline — cum = eigene kumulierte Punkte über alle gefahrenen")
    o.append("// Etappen (anders als OPTIMAL_TEAMS, das je Etappe wechseln kann).")
    if kanter_kader:
        team_js = js_str(kanter_kader["team"]) if kanter_kader["team"] else "null"
        o.append(
            "export const KANTER_KADER = {{stage:{stage}, team:{team}, teamCost:{tc}, teamPoints:{tp}, "
            "totalCost:{totc}, totalPoints:{totp}, riders:[{riders}], cum:[{cum}]}};".format(
                stage=kanter_kader["stage"], team=team_js, tc=kanter_kader["teamCost"],
                tp=kanter_kader["teamPoints"], totc=kanter_kader["totalCost"], totp=kanter_kader["totalPoints"],
                riders=riders_js_array(kanter_kader["riders"]), cum=",".join(map(str, kanter_kader["cum"])),
            )
        )
    else:
        o.append("export const KANTER_KADER = null;")
    o.append("")

    o.append("// Teamfunk — Trash Talk der Runde (Sheet 'Quotes' in der xlsx pflegen).")
    o.append("export const QUOTES = [")
    for q in quotes:
        o.append(f"  {{by:{js_str(q['by'])}, txt:{js_str(q['txt'])}}},")
    o.append("];")
    o.append("")

    o.append("""export function buildPlayers(){
  return PLAYERS.map(p => {
    const stages = p.stages;
    const total = stages.reduce((a, b) => a + b, 0);
    const cum = []; let s = 0;
    stages.forEach(v => { s += v; cum.push(s); });
    return { ...p, total, cum };
  });
}
""")

    OUT.write_text("\n".join(o), encoding="utf-8")


if __name__ == "__main__":
    build()
